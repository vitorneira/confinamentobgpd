# -*- coding: utf-8 -*-
"""
Gera a planilha-modelo de IMPORTAÇÃO do sistema de confinamento.
Duas abas: Pesagens e Cadastro_Animais. Com instruções, linha de exemplo,
validações (data, número) e listas suspensas (fazenda, categoria, curral).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

FONT = "Arial"
hdr_fill = PatternFill("solid", fgColor="2E5E3A")      # verde escuro
hdr_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
title_font = Font(name=FONT, bold=True, size=14, color="2E5E3A")
note_font = Font(name=FONT, italic=True, size=9, color="666666")
ex_font = Font(name=FONT, italic=True, color="1F5FBF")  # exemplo em azul
norm = Font(name=FONT, size=10)
thin = Side(style="thin", color="C9D6CC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ---- listas de referência (por fazenda) — o sistema valida contra o cadastro real ----
FAZENDAS = ["BG", "PD"]
CATEGORIAS = ["Touro Bonsmara", "Touro Nelore", "Bezerro", "Novilha Nelore",
              "Touro Bonsmara Descarte", "Boi", "F Cruz. Industrial"]
CURRAIS = ["1","2","3","4","5","6","7","8","9","10","4-TN","(a definir)"]
TIPO_ENTRADA = ["individual", "agregado"]

def style_header(ws, row, headers, widths):
    for c,(h,w) in enumerate(zip(headers,widths), start=1):
        cell = ws.cell(row, c, h)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center; cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

def add_list_dv(ws, col_letter, options, first=None, last=1000):
    dv = DataValidation(type="list", formula1='"'+",".join(options)+'"', allow_blank=True)
    dv.error = "Escolha um valor da lista."; dv.errorTitle = "Valor inválido"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first}:{col_letter}{last}")

def add_num_dv(ws, col_letter, first, last=1000, mn=0):
    dv = DataValidation(type="decimal", operator="greaterThan", formula1=str(mn), allow_blank=True)
    dv.error = "Digite um número maior que %s." % mn; dv.errorTitle = "Número inválido"
    ws.add_data_validation(dv); dv.add(f"{col_letter}{first}:{col_letter}{last}")

wb = Workbook()

# =========================================================
# ABA 1 — PESAGENS
# =========================================================
ws = wb.active; ws.title = "Pesagens"
ws["A1"] = "MODELO DE IMPORTAÇÃO — PESAGENS"; ws["A1"].font = title_font
ws["A2"] = ("Uma linha por pesagem. Preencha a partir da linha 5. Não altere os títulos da linha 4. "
            "Data no formato AAAA-MM-DD. Peso em kg (número). Categoria e Curral devem existir no cadastro da fazenda.")
ws["A2"].font = note_font; ws.merge_cells("A2:F2"); ws["A2"].alignment = left
ws.row_dimensions[2].height = 30

headers = ["Fazenda", "Data", "Brinco", "Categoria", "Curral", "Peso (kg)"]
widths  = [12, 14, 18, 24, 12, 12]
style_header(ws, 4, headers, widths)

# comentários de ajuda no cabeçalho
ws["C4"].comment = Comment("Para lote AGREGADO (sem brinco), deixe em branco e informe o lote na coluna adicional? "
                          "Não: agregado é pesado no Cadastro. Aqui é 1 brinco por linha.", "Sistema")
ws["B4"].comment = Comment("Formato AAAA-MM-DD, ex.: 2026-08-15", "Sistema")

# linha de exemplo (azul, itálico)
exemplo = ["BG", "2026-08-15", "BBG 3339", "Touro Bonsmara", "2", 548]
for c,v in enumerate(exemplo, start=1):
    cell = ws.cell(5, c, v); cell.font = ex_font; cell.border = border
    cell.alignment = center if c!=4 else left
ws["A6"] = "↑ linha de exemplo (pode apagar)"; ws["A6"].font = note_font

# validações
add_list_dv(ws, "A", FAZENDAS, first=5)
add_list_dv(ws, "D", CATEGORIAS, first=5)
add_list_dv(ws, "E", CURRAIS, first=5)
add_num_dv(ws, "F", first=5, mn=0)
# bordas na área de digitação
for r in range(7, 205):
    for c in range(1,7): ws.cell(r,c).border = border; ws.cell(r,c).font = norm
ws.freeze_panes = "A5"

# =========================================================
# ABA 2 — CADASTRO DE ANIMAIS (entrada de lote novo)
# =========================================================
ws2 = wb.create_sheet("Cadastro_Animais")
ws2["A1"] = "MODELO DE IMPORTAÇÃO — CADASTRO DE ANIMAIS (ENTRADA)"; ws2["A1"].font = title_font
ws2["A2"] = ("Cadastro de animais que ENTRAM no confinamento. Dois modos:\n"
             "• INDIVIDUAL: uma linha por animal, com Brinco e Peso de Entrada.\n"
             "• AGREGADO: uma linha por LOTE, sem brinco, informando Quantidade e Peso Médio de Entrada.\n"
             "Preencha a partir da linha 6. Data AAAA-MM-DD. Não altere os títulos da linha 5.")
ws2["A2"].font = note_font; ws2.merge_cells("A2:I2"); ws2["A2"].alignment = left
ws2.row_dimensions[2].height = 62

headers2 = ["Fazenda","Tipo Entrada","Data Entrada","Brinco","Categoria","Curral",
            "Lote/Origem","Qtd (agregado)","Peso Entrada (kg)"]
widths2  = [12,14,14,18,22,10,18,14,16]
style_header(ws2, 5, headers2, widths2)

ws2["B5"].comment = Comment("individual = 1 animal por linha (com brinco).\nagregado = 1 lote por linha (sem brinco, usa Qtd + Peso Médio).", "Sistema")
ws2["D5"].comment = Comment("Só para tipo 'individual'. Deixe vazio no 'agregado'.", "Sistema")
ws2["H5"].comment = Comment("Só para tipo 'agregado': número de cabeças do lote.", "Sistema")
ws2["I5"].comment = Comment("Individual: peso do animal.\nAgregado: peso MÉDIO do lote.", "Sistema")

# dois exemplos: individual e agregado
ex_ind = ["BG","individual","2026-06-30","BBG 3339","Touro Bonsmara","2","Retiro 1","", 670]
ex_agg = ["PD","agregado","2026-07-17","","Boi","4","Compra Fulano", 90, 385]
for r,ex in [(6,ex_ind),(7,ex_agg)]:
    for c,v in enumerate(ex, start=1):
        cell = ws2.cell(r,c,v); cell.font = ex_font; cell.border = border
        cell.alignment = center if c not in (5,7) else left
ws2["A8"] = "↑ duas linhas de exemplo (individual e agregado — pode apagar)"; ws2["A8"].font = note_font

add_list_dv(ws2, "A", FAZENDAS, first=6)
add_list_dv(ws2, "B", TIPO_ENTRADA, first=6)
add_list_dv(ws2, "E", CATEGORIAS, first=6)
add_list_dv(ws2, "F", CURRAIS, first=6)
add_num_dv(ws2, "H", first=6, mn=0)
add_num_dv(ws2, "I", first=6, mn=0)
for r in range(9, 209):
    for c in range(1,10): ws2.cell(r,c).border = border; ws2.cell(r,c).font = norm
ws2.freeze_panes = "A6"

# =========================================================
# ABA 3 — LEIA-ME
# =========================================================
ws3 = wb.create_sheet("LEIA-ME", 0)
ws3["A1"] = "COMO USAR ESTE MODELO"; ws3["A1"].font = title_font
linhas = [
 "",
 "Este arquivo é o formato que o SISTEMA aceita para importar dados.",
 "Fluxo: o campo preenche a FOLHA DE CAMPO (papel) → você digitaliza (PDF) →",
 "o Cowork converte para este modelo → você sobe este arquivo no sistema.",
 "",
 "ABAS:",
 "• Pesagens — uma linha por pesagem (individual, com brinco).",
 "• Cadastro_Animais — entrada de animais novos (individual ou agregado).",
 "",
 "REGRAS:",
 "• Não altere os nomes dos títulos das colunas.",
 "• Datas sempre no formato AAAA-MM-DD (ex.: 2026-08-15).",
 "• Fazenda, Categoria e Curral: use as listas suspensas.",
 "• Células de exemplo (em azul) podem ser apagadas antes de subir.",
 "• Pesos e quantidades: apenas números.",
 "",
 "As listas de Categoria e Curral neste modelo são uma referência inicial.",
 "O sistema valida contra o cadastro real de cada fazenda no momento da importação.",
]
for i,t in enumerate(linhas, start=2):
    ws3.cell(i,1,t).font = norm if not t.endswith(":") else Font(name=FONT,bold=True,size=10)
ws3.column_dimensions["A"].width = 95

import os
out = "/home/claude/pacote-confinamento/Modelo_Importacao_Confinamento.xlsx"
wb.save(out)
print("salvo:", out)
